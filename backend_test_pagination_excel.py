"""
Backend tests for two new features:
1. Pagination on list endpoints (skip/limit clamping & defaults).
2. Excel catalog import (template download + xlsx upload: structured, AI fallback, validation).
"""
import base64
import io
import json
import os
import time
import uuid

import requests
from openpyxl import Workbook

BASE_URL = "https://pharma-checkout-8.preview.emergentagent.com/api"
TIMEOUT = 30

PHARMACY_PHONE = "07700000001"
PHARMACY_PASS = "pass123"
SUPPLIER_PHONE = "07811111111"
SUPPLIER_PASS = "sup1"
ADMIN_PHONE = "0000000000"
ADMIN_PASS = "admin123"

results: list[tuple[str, bool, str]] = []


def record(name: str, ok: bool, detail: str = ""):
    results.append((name, ok, detail))
    icon = "PASS" if ok else "FAIL"
    print(f"[{icon}] {name}{(' :: ' + detail) if detail else ''}")


def login(phone: str, password: str) -> tuple[str | None, dict | None]:
    r = requests.post(f"{BASE_URL}/auth/login", json={"phone": phone, "password": password}, timeout=TIMEOUT)
    if r.status_code != 200:
        return None, {"status": r.status_code, "body": r.text[:200]}
    j = r.json()
    return j.get("token"), j


def auth(tok: str) -> dict:
    return {"Authorization": f"Bearer {tok}"}


def ensure_supplier_enabled(admin_tok: str, supplier_id: str) -> bool:
    r = requests.patch(
        f"{BASE_URL}/admin/users/supplier/{supplier_id}",
        headers=auth(admin_tok),
        json={"disabled": False},
        timeout=TIMEOUT,
    )
    return r.status_code == 200


# ---------------------------------------------------------------- LOGINS
print("\n=== Login phase ===")
admin_tok, admin_info = login(ADMIN_PHONE, ADMIN_PASS)
if not admin_tok:
    record("admin_login", False, str(admin_info))
    raise SystemExit("Cannot login admin – aborting.")
record("admin_login", True, "0000000000/admin123")

pharmacy_tok, pharmacy_info = login(PHARMACY_PHONE, PHARMACY_PASS)
if not pharmacy_tok:
    record("pharmacy_login", False, str(pharmacy_info))
    raise SystemExit("Cannot login pharmacy – aborting.")
record("pharmacy_login", True, "07700000001/pass123")

supplier_tok, supplier_info = login(SUPPLIER_PHONE, SUPPLIER_PASS)
if not supplier_tok:
    # Try to find supplier by listing admin users and re-enable
    r = requests.get(f"{BASE_URL}/admin/users?role=supplier", headers=auth(admin_tok), timeout=TIMEOUT)
    found = None
    if r.status_code == 200:
        for u in r.json():
            if u.get("phone") == SUPPLIER_PHONE:
                found = u
                break
    if found:
        ok = ensure_supplier_enabled(admin_tok, found["id"])
        record("supplier_reenable_via_admin", ok, f"id={found['id']}")
        supplier_tok, supplier_info = login(SUPPLIER_PHONE, SUPPLIER_PASS)
    if not supplier_tok:
        record("supplier_login", False, str(supplier_info))
        raise SystemExit("Cannot login supplier – aborting.")
record("supplier_login", True, "07811111111/sup1")

# Make sure pharmacy + supplier each have at least a few records so pagination is meaningful
def seed_medicines_if_needed():
    r = requests.get(f"{BASE_URL}/medicines", headers=auth(pharmacy_tok), timeout=TIMEOUT)
    cnt = len(r.json()) if r.status_code == 200 else 0
    needed = max(0, 4 - cnt)
    for i in range(needed):
        requests.post(
            f"{BASE_URL}/medicines",
            headers=auth(pharmacy_tok),
            json={"name": f"PgTestMed{i}-{uuid.uuid4().hex[:6]}", "quantity": 5, "price": 1000 + i},
            timeout=TIMEOUT,
        )

def seed_orders_if_needed():
    r = requests.get(f"{BASE_URL}/orders", headers=auth(pharmacy_tok), timeout=TIMEOUT)
    cnt = len(r.json()) if r.status_code == 200 else 0
    needed = max(0, 3 - cnt)
    for i in range(needed):
        requests.post(
            f"{BASE_URL}/orders",
            headers=auth(pharmacy_tok),
            json={"items": [{"name": f"OrderItemPg{i}", "quantity": 2}]},
            timeout=TIMEOUT,
        )

def seed_supplier_products_if_needed():
    r = requests.get(f"{BASE_URL}/supplier/products", headers=auth(supplier_tok), timeout=TIMEOUT)
    cnt = len(r.json()) if r.status_code == 200 else 0
    needed = max(0, 4 - cnt)
    for i in range(needed):
        requests.post(
            f"{BASE_URL}/supplier/products",
            headers=auth(supplier_tok),
            json={"name": f"PgTestProd{i}-{uuid.uuid4().hex[:6]}", "quantity": 50, "price": 2000 + i * 10},
            timeout=TIMEOUT,
        )

print("\n=== Seeding (idempotent) ===")
seed_medicines_if_needed()
seed_orders_if_needed()
seed_supplier_products_if_needed()


# ---------------------------------------------------------------- PAGINATION TESTS
print("\n=== Pagination tests ===")

ENDPOINTS = [
    # (name, url_path, token, default_limit, response_kind)
    ("/medicines", "/medicines", pharmacy_tok, 200, "list"),
    ("/orders", "/orders", pharmacy_tok, 100, "list"),
    ("/supplier/products", "/supplier/products", supplier_tok, 200, "list"),
    ("/admin/users", "/admin/users", admin_tok, 200, "list"),
    ("/admin/orders", "/admin/orders", admin_tok, 200, "list"),
    ("/admin/products", "/admin/products", admin_tok, 200, "list"),
    ("/admin/commissions", "/admin/commissions", admin_tok, 200, "records_obj"),
    ("/supplier/commissions", "/supplier/commissions", supplier_tok, 200, "records_obj"),
]


def get_list(resp_json, kind):
    if kind == "list":
        return resp_json if isinstance(resp_json, list) else None
    if kind == "records_obj":
        return resp_json.get("records") if isinstance(resp_json, dict) else None
    return None


for name, path, tok, default_limit, kind in ENDPOINTS:
    headers = auth(tok)

    # (a) baseline (no params)
    r = requests.get(f"{BASE_URL}{path}", headers=headers, timeout=TIMEOUT)
    if r.status_code != 200:
        record(f"PAG {name} baseline", False, f"status={r.status_code} body={r.text[:160]}")
        continue
    base = get_list(r.json(), kind)
    if base is None:
        record(f"PAG {name} baseline", False, "Could not extract records list")
        continue
    record(f"PAG {name} baseline", True, f"count={len(base)}")
    base_count = len(base)

    # (b) limit=1
    r = requests.get(f"{BASE_URL}{path}", headers=headers, params={"limit": 1}, timeout=TIMEOUT)
    arr = get_list(r.json(), kind) if r.status_code == 200 else None
    if r.status_code != 200 or arr is None:
        record(f"PAG {name} limit=1", False, f"status={r.status_code}")
    else:
        ok = (len(arr) == 1) if base_count >= 1 else (len(arr) == 0)
        record(f"PAG {name} limit=1", ok, f"len={len(arr)} (base={base_count})")

    # (c) skip=1 limit=2 -> should match base[1:3] when base has enough
    r = requests.get(f"{BASE_URL}{path}", headers=headers, params={"skip": 1, "limit": 2}, timeout=TIMEOUT)
    arr = get_list(r.json(), kind) if r.status_code == 200 else None
    if r.status_code != 200 or arr is None:
        record(f"PAG {name} skip=1&limit=2", False, f"status={r.status_code}")
    else:
        # cannot strictly compare ids across endpoints (admin/users, admin/products mix two collections),
        # but we can verify length doesn't exceed 2.
        ok_len = len(arr) <= 2
        record(f"PAG {name} skip=1&limit=2", ok_len, f"len={len(arr)}")

    # (d) limit=999 -> capped at 500 (no error)
    r = requests.get(f"{BASE_URL}{path}", headers=headers, params={"limit": 999}, timeout=TIMEOUT)
    arr = get_list(r.json(), kind) if r.status_code == 200 else None
    if r.status_code != 200 or arr is None:
        record(f"PAG {name} limit=999", False, f"status={r.status_code} body={r.text[:160]}")
    else:
        record(f"PAG {name} limit=999", len(arr) <= 500, f"len={len(arr)} <=500")

    # (e) limit=0 and limit=-5 -> fall back to default, no crash
    for bad in (0, -5):
        r = requests.get(f"{BASE_URL}{path}", headers=headers, params={"limit": bad}, timeout=TIMEOUT)
        arr = get_list(r.json(), kind) if r.status_code == 200 else None
        ok = (r.status_code == 200) and (arr is not None)
        # Length should not exceed default_limit (cannot guarantee it equals base_count because of clamp logic)
        if ok:
            ok = len(arr) <= max(default_limit, base_count)
        record(f"PAG {name} limit={bad}", ok, f"status={r.status_code} len={len(arr) if arr is not None else 'NA'}")

    # (f) skip=99999 -> empty list, no crash
    r = requests.get(f"{BASE_URL}{path}", headers=headers, params={"skip": 99999}, timeout=TIMEOUT)
    arr = get_list(r.json(), kind) if r.status_code == 200 else None
    if r.status_code != 200 or arr is None:
        record(f"PAG {name} skip=99999", False, f"status={r.status_code}")
    else:
        record(f"PAG {name} skip=99999", len(arr) == 0, f"len={len(arr)}")


# Verify supplier/commissions returns extra fields and they remain consistent
print("\n=== /supplier/commissions structure (full dataset based) ===")
r = requests.get(f"{BASE_URL}/supplier/commissions", headers=auth(supplier_tok),
                 params={"limit": 1}, timeout=TIMEOUT)
if r.status_code == 200:
    j = r.json()
    keys = set(j.keys())
    needed = {"records", "monthly", "outstanding", "total_due", "total_sales", "rate"}
    ok = needed.issubset(keys)
    record("supplier_commissions_keys", ok, f"missing={needed - keys}")
    record("supplier_commissions_records_lim1", isinstance(j.get("records"), list) and len(j["records"]) <= 1,
           f"records_len={len(j.get('records', []))}")
    # monthly should be from full dataset; pagination does NOT zero them out
    record("supplier_commissions_monthly_independent",
           isinstance(j.get("monthly"), list),
           f"monthly_len={len(j.get('monthly') or [])}")
else:
    record("supplier_commissions_struct", False, f"status={r.status_code}")

print("\n=== /admin/commissions structure ===")
r = requests.get(f"{BASE_URL}/admin/commissions", headers=auth(admin_tok),
                 params={"limit": 1}, timeout=TIMEOUT)
if r.status_code == 200:
    j = r.json()
    record("admin_commissions_keys", "records" in j and "stats" in j, f"keys={list(j.keys())}")
    record("admin_commissions_records_lim1",
           isinstance(j.get("records"), list) and len(j["records"]) <= 1,
           f"records_len={len(j.get('records', []))}")
else:
    record("admin_commissions_struct", False, f"status={r.status_code}")


# ---------------------------------------------------------------- EXCEL CATALOG IMPORT
print("\n=== Excel catalog import tests ===")

# (b) GET template
r = requests.get(f"{BASE_URL}/supplier/catalog/template", headers=auth(supplier_tok), timeout=TIMEOUT)
ok_status = r.status_code == 200
ok_ctype = "spreadsheetml" in r.headers.get("content-type", "")
ok_size = len(r.content) > 1000
ok_pk = r.content[:2] == b"PK"
record("template_status_200", ok_status, f"status={r.status_code}")
record("template_content_type", ok_ctype, f"ct={r.headers.get('content-type')}")
record("template_size_gt_1000", ok_size, f"size={len(r.content)}")
record("template_zip_signature_PK", ok_pk, f"first2={r.content[:2]!r}")

template_bytes = r.content if ok_status else b""

# Role enforcement on /template (pharmacy should get 403)
r2 = requests.get(f"{BASE_URL}/supplier/catalog/template", headers=auth(pharmacy_tok), timeout=TIMEOUT)
record("template_role_enforcement", r2.status_code == 403, f"pharmacy_status={r2.status_code}")


def upload_xlsx(blob: bytes, filename: str) -> str | None:
    payload = {
        "file_b64": base64.b64encode(blob).decode("ascii"),
        "file_type": "xlsx",
        "filename": filename,
    }
    r = requests.post(f"{BASE_URL}/supplier/catalog/upload",
                      headers=auth(supplier_tok), json=payload, timeout=TIMEOUT)
    if r.status_code != 200:
        print(f"   upload failed: status={r.status_code} body={r.text[:200]}")
        return None
    return r.json().get("job_id")


def poll_job(job_id: str, max_seconds: int = 30) -> dict | None:
    for _ in range(max_seconds):
        r = requests.get(f"{BASE_URL}/supplier/catalog/jobs/{job_id}",
                         headers=auth(supplier_tok), timeout=TIMEOUT)
        if r.status_code == 200:
            data = r.json()
            job = data.get("job", {})
            if job.get("status") in ("review", "failed", "published"):
                return data
        time.sleep(1)
    return None


# (c+d) Upload the template itself -> structured parsing -> 3 rows
if template_bytes:
    job_id = upload_xlsx(template_bytes, "test_template.xlsx")
    record("structured_upload_jobid", bool(job_id), f"job_id={job_id}")
    if job_id:
        result = poll_job(job_id, max_seconds=30)
        if not result:
            record("structured_upload_terminal", False, "did not reach review/failed in 30s")
        else:
            job = result["job"]
            items = result.get("items", [])
            record("structured_status_review", job.get("status") == "review",
                   f"status={job.get('status')} method={job.get('extraction_method')}")
            record("structured_method", job.get("extraction_method") == "excel_structured",
                   f"method={job.get('extraction_method')}")
            record("structured_total_items_3", job.get("total_items") == 3,
                   f"total_items={job.get('total_items')}")
            # Validate items structure
            ok_items = (
                isinstance(items, list) and len(items) == 3
                and all((it.get("extracted") or {}).get("name") for it in items)
                and all(float((it.get("extracted") or {}).get("price") or 0) > 0 for it in items)
                and all((it.get("extracted") or {}).get("category") for it in items)
            )
            record("structured_items_have_name_price_cat", ok_items,
                   f"n_items={len(items)}")


# (e) Bad headers -> AI fallback (or 0 items, both acceptable)
def build_bad_headers_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Garbage"
    ws.append(["foo", "bar", "baz"])
    ws.append(["alpha", 123, "x"])
    ws.append(["beta", 456, "y"])
    ws2 = wb.create_sheet("Other")
    ws2.append(["who", "knows", "what"])
    ws2.append(["a", 1, "b"])
    ws2.append(["c", 2, "d"])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

bad_blob = build_bad_headers_xlsx()
job_id = upload_xlsx(bad_blob, "bad_headers.xlsx")
record("bad_headers_upload_jobid", bool(job_id), f"job_id={job_id}")
if job_id:
    result = poll_job(job_id, max_seconds=45)
    if not result:
        record("bad_headers_terminal", False, "did not reach review/failed in 45s")
    else:
        job = result["job"]
        method = job.get("extraction_method")
        status_ = job.get("status")
        total = job.get("total_items", 0)
        # Acceptable: ai_fallback OR review with 0 items
        ok = (
            (method == "excel_ai_fallback" and status_ in ("review", "failed"))
            or (status_ == "review" and total == 0)
        )
        record("bad_headers_graceful", ok,
               f"status={status_} method={method} total_items={total}")


# (f) Valid headers but invalid prices -> rejected_invalid > 0
def build_invalid_prices_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"
    ws.append(["product_name", "price", "quantity", "category"])
    ws.append(["BadDrugA", "abc", 5, "test"])      # non-numeric price -> 0 -> rejected
    ws.append(["BadDrugB", 0, 5, "test"])          # zero -> rejected
    ws.append(["BadDrugC", -100, 5, "test"])       # negative -> rejected
    ws.append(["GoodDrugD", 1500, 20, "test"])     # valid -> kept
    ws.append(["", 999, 1, "test"])                # empty name -> rejected
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

inv_blob = build_invalid_prices_xlsx()
job_id = upload_xlsx(inv_blob, "invalid_prices.xlsx")
record("invalid_prices_upload_jobid", bool(job_id), f"job_id={job_id}")
if job_id:
    result = poll_job(job_id, max_seconds=30)
    if not result:
        record("invalid_prices_terminal", False, "did not reach review/failed in 30s")
    else:
        job = result["job"]
        items = result.get("items", [])
        status_ = job.get("status")
        rejected = job.get("rejected_invalid", 0)
        total = job.get("total_items", 0)
        record("invalid_prices_status_review", status_ == "review", f"status={status_}")
        record("invalid_prices_rejected_gt_0", (rejected or 0) > 0,
               f"rejected_invalid={rejected}")
        record("invalid_prices_total_only_valid", total == 1,
               f"total_items={total} (expected 1, GoodDrugD)")
        if items:
            names = [(it.get("extracted") or {}).get("name") for it in items]
            record("invalid_prices_only_good_kept",
                   all(n == "GoodDrugD" for n in names),
                   f"names={names}")


# ---------------------------------------------------------------- SUMMARY
print("\n" + "=" * 70)
passed = sum(1 for _, ok, _ in results if ok)
failed = sum(1 for _, ok, _ in results if not ok)
print(f"TOTAL: {passed} passed, {failed} failed (out of {len(results)})")
print("=" * 70)
if failed:
    print("\nFAILURES:")
    for n, ok, d in results:
        if not ok:
            print(f"  - {n}: {d}")
