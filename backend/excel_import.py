"""
Excel / CSV catalog import for supplier catalogs.

Feature scope (per user's explicit brief):
  * NEW endpoints under /api/orders/excel/* — read the uploaded file,
    intelligently map its columns (Arabic + English variants), and
    return parsed rows for review.
  * On commit, every row goes through the EXISTING `_batches.create_batch`
    used by /medicines/buy-v2 — so FIFO, batch expiry, stock mirror,
    return credits, expiry alerts all continue to work identically.
  * DOES NOT modify OCR, purchase, sales, inventory, FIFO, batches
    or permissions. Only ADDS.
"""
from __future__ import annotations

import base64
import csv
import io
import logging
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

_db = None
router_excel = APIRouter(prefix="/api", tags=["excel_import"])


def init(db_instance) -> None:
    global _db
    _db = db_instance


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------------- SMART COLUMN MAPPER ---------------------------------

def _norm_header(s: Any) -> str:
    """Normalize a header cell for tolerant matching:
    - lowercase, strip diacritics + non-alnum
    - unify Arabic variants (ا/أ/إ/آ → ا, ة → ه, ى → ي)
    """
    if s is None:
        return ""
    v = str(s).strip().lower()
    # Unify Arabic normalization
    v = v.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    v = v.replace("ة", "ه").replace("ى", "ي").replace("ؤ", "و").replace("ئ", "ي")
    # Remove diacritics range
    v = re.sub(r"[\u064b-\u065f\u0670]", "", v)
    # Keep only alnum + arabic letters, then squash
    v = re.sub(r"[^0-9a-z\u0600-\u06ff]+", "", v)
    return v


# Field → list of accepted header aliases (all normalized)
FIELD_ALIASES: Dict[str, List[str]] = {
    "barcode": [
        "barcode", "bar", "ean", "gtin", "code", "scancode", "sku",
        "الباركود", "باركود", "رمز", "رمزالمنتج", "بار", "الرمز",
    ],
    "name": [
        "name", "medicine", "medicinename", "product", "productname",
        "drug", "drugname", "item", "itemname", "description", "desc",
        "الاسم", "اسم", "اسمالدواء", "اسمالمنتج", "دواء", "المنتج",
        "الصنف", "صنف", "التسميه", "التسميةالتجاريه", "الاسمالتجاري",
    ],
    "strength": [
        "strength", "dose", "dosage", "concentration", "mg", "ml",
        "التركيز", "تركيز", "الجرعه", "جرعه",
    ],
    "dosage_form": [
        "form", "dosageform", "pharmaform", "type",
        "الشكل", "شكل", "الشكلالصيدلاني", "الفئه", "نوع",
    ],
    "quantity": [
        "qty", "quantity", "stock", "count", "units", "packs", "amount",
        "الكميه", "كميه", "الرصيد", "رصيد", "العدد", "عدد",
    ],
    "purchase_price": [
        "price", "cost", "purchaseprice", "buyprice", "unitprice",
        "wholesale", "wholesaleprice",
        "السعر", "سعر", "سعرالشراء", "التكلفه", "تكلفه", "الجمله", "الكلفه",
    ],
    "expiry_date": [
        "expiry", "expirydate", "exp", "expdate", "expiration",
        "الصلاحيه", "الصلاحية", "الانتهاء", "تاريخالانتهاء", "انتهاء", "تاريخالصلاحيه",
    ],
    "batch_number": [
        "batch", "batchno", "batchnumber", "lot", "lotno", "lotnumber",
        "التشغيله", "تشغيله", "رقمالتشغيله", "لوت", "الدفعه",
    ],
    "manufacturer": [
        "manufacturer", "maker", "brand", "company", "vendor", "producer",
        "الشركه", "شركه", "المصنع", "المنتج", "منشا", "الشركهالمصنعه",
    ],
}


def _map_columns(headers: List[Any]) -> Dict[str, int]:
    """Return {field_name: column_index} — best match by alias list.
    Ignores empty/None headers."""
    mapping: Dict[str, int] = {}
    normed = [_norm_header(h) for h in headers]
    for field, aliases in FIELD_ALIASES.items():
        best_idx = -1
        for idx, h in enumerate(normed):
            if not h:
                continue
            # exact alias match
            if h in aliases:
                best_idx = idx
                break
            # substring / contained match — good for headers like
            # "Medicine Name (ar)" → normalized "medicinenamear"
            for a in aliases:
                if len(a) >= 3 and (a in h or h in a):
                    best_idx = idx
                    break
            if best_idx != -1:
                break
        if best_idx != -1:
            mapping[field] = best_idx
    return mapping


def _parse_number(v: Any) -> float:
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def _parse_int(v: Any) -> int:
    return int(round(_parse_number(v)))


def _parse_expiry(v: Any) -> Optional[str]:
    if not v:
        return None
    try:
        # openpyxl may return datetime already
        if hasattr(v, "isoformat"):
            return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
    except Exception:
        pass
    s = str(v).strip()
    if not s:
        return None
    # YYYY-MM-DD
    m = re.match(r"^(\d{4})[-/](\d{1,2})(?:[-/](\d{1,2}))?$", s)
    if m:
        y, mo, d = m.group(1), m.group(2).zfill(2), (m.group(3) or "01").zfill(2)
        return f"{y}-{mo}-{d}"
    # DD/MM/YYYY or DD-MM-YYYY
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})$", s)
    if m:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        if len(y) == 2:
            y = "20" + y
        return f"{y}-{mo}-{d}"
    # MM/YYYY
    m = re.match(r"^(\d{1,2})[-/](\d{4})$", s)
    if m:
        mo, y = m.group(1).zfill(2), m.group(2)
        return f"{y}-{mo}-01"
    return None


# --------------------------- FILE PARSER ------------------------------

def _decode_bytes(b64: str) -> bytes:
    try:
        return base64.b64decode(b64, validate=False)
    except Exception:
        raise HTTPException(400, "الملف غير صالح")


def _parse_csv(raw: bytes) -> Tuple[List[str], List[List[Any]]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    # Sniff delimiter
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delim = dialect.delimiter
    except Exception:
        delim = ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return [], []
    headers = rows[0]
    return headers, rows[1:]


def _parse_xlsx(raw: bytes) -> Tuple[List[str], List[List[Any]]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return [], []
    data: List[List[Any]] = []
    for r in rows_iter:
        # Skip fully-empty rows
        if not any(c is not None and str(c).strip() != "" for c in r):
            continue
        data.append(list(r))
        if len(data) > 30000:                        # hard cap
            break
    return headers, data


def _parse_xls(raw: bytes) -> Tuple[List[str], List[List[Any]]]:
    # xlrd 2.0 dropped xls; fall back to a soft error
    try:
        import xlrd  # type: ignore
    except Exception:
        raise HTTPException(400, "الرجاء تحويل الملف إلى XLSX أو CSV")
    book = xlrd.open_workbook(file_contents=raw)
    sh = book.sheet_by_index(0)
    if sh.nrows == 0:
        return [], []
    headers = [sh.cell_value(0, c) for c in range(sh.ncols)]
    data = []
    for r in range(1, sh.nrows):
        row = [sh.cell_value(r, c) for c in range(sh.ncols)]
        if not any(str(c).strip() for c in row):
            continue
        data.append(row)
    return headers, data


def _parse_file(filename: str, raw: bytes) -> Tuple[List[str], List[List[Any]]]:
    fn = (filename or "").lower()
    if fn.endswith(".csv") or fn.endswith(".txt"):
        return _parse_csv(raw)
    if fn.endswith(".xlsx"):
        return _parse_xlsx(raw)
    if fn.endswith(".xls"):
        return _parse_xls(raw)
    # Heuristic on magic bytes
    if raw[:4] == b"PK\x03\x04":                # xlsx zip signature
        return _parse_xlsx(raw)
    return _parse_csv(raw)


# ============================= MODELS =================================

class PreviewIn(BaseModel):
    filename: str
    file_base64: str


class ExcelItemIn(BaseModel):
    name: str
    quantity: int
    purchase_price: float
    selling_price: Optional[float] = None
    barcode: Optional[str] = None
    expiry_date: Optional[str] = None
    batch_number: Optional[str] = None
    manufacturer: Optional[str] = None
    strength: Optional[str] = None
    dosage_form: Optional[str] = None


class CommitIn(BaseModel):
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None
    items: List[ExcelItemIn]


# ============================= ROUTES =================================

def install_routes(require_role):
    router_excel.routes.clear()

    @router_excel.post("/orders/excel/preview")
    async def _preview(data: PreviewIn,
                        user: dict = Depends(require_role("pharmacy"))):
        """Parse an uploaded catalog and return rows for the review screen."""
        b64 = (data.file_base64 or "").strip()
        if b64.startswith("data:") and "," in b64:
            b64 = b64.split(",", 1)[1]
        if len(b64) < 20:
            raise HTTPException(400, "الملف مفقود")
        raw = _decode_bytes(b64)
        if len(raw) > 30 * 1024 * 1024:
            raise HTTPException(400, "الملف كبير جداً (الحد الأقصى 30MB)")
        headers, rows = _parse_file(data.filename or "", raw)
        if not headers:
            raise HTTPException(400, "لم يتم قراءة أي صف من الملف")
        mapping = _map_columns(headers)
        if "name" not in mapping:
            raise HTTPException(400,
                "لم نتمكن من تحديد عمود اسم الدواء — تأكد أن أول صف يحتوي على العناوين")

        def get(row: List[Any], key: str, default=None):
            i = mapping.get(key)
            if i is None or i >= len(row):
                return default
            v = row[i]
            return None if (v is None or str(v).strip() == "") else v

        items: List[Dict[str, Any]] = []
        for r in rows:
            name = get(r, "name")
            if not name:
                continue
            items.append({
                "name": str(name).strip(),
                "barcode": (str(get(r, "barcode") or "").strip() or None),
                "quantity": _parse_int(get(r, "quantity", 0)),
                "purchase_price": round(_parse_number(get(r, "purchase_price", 0)), 2),
                "expiry_date": _parse_expiry(get(r, "expiry_date")),
                "batch_number": (str(get(r, "batch_number") or "").strip() or None),
                "manufacturer": (str(get(r, "manufacturer") or "").strip() or None),
                "strength": (str(get(r, "strength") or "").strip() or None),
                "dosage_form": (str(get(r, "dosage_form") or "").strip() or None),
            })
        return {
            "items": items,
            "count": len(items),
            "columns_detected": mapping,
            "headers": [str(h or "") for h in headers],
        }

    @router_excel.post("/orders/excel/commit", status_code=201)
    async def _commit(data: CommitIn,
                       user: dict = Depends(require_role("pharmacy"))):
        """Persist reviewed items via the SAME batch pipeline used by
        /medicines/buy-v2 → FIFO + batch expiry + stock mirror + supplier
        debts all continue to work identically."""
        if not data.items:
            raise HTTPException(400, "لا يوجد أصناف للحفظ")
        import batches as _batches

        pharmacy_id = user["sub"]
        new_count = 0
        updated_count = 0
        errors: List[Dict[str, Any]] = []
        succeeded: List[Dict[str, Any]] = []

        for idx, it in enumerate(data.items):
            try:
                name = (it.name or "").strip()
                if not name:
                    raise ValueError("الاسم مفقود")
                qty = int(it.quantity or 0)
                if qty <= 0:
                    raise ValueError("الكمية غير صالحة")

                purchase_price = float(it.purchase_price or 0)
                selling_price = float(it.selling_price or 0) if it.selling_price \
                    else round(purchase_price * 1.25, 2)
                expiry_iso = _parse_expiry(it.expiry_date)

                q: Dict[str, Any] = {"pharmacy_id": pharmacy_id}
                existing = None
                if it.barcode:
                    existing = await _db.medicines.find_one(
                        {**q, "barcode": it.barcode}, {"_id": 0})
                if not existing:
                    existing = await _db.medicines.find_one(
                        {**q, "name": name}, {"_id": 0})

                if existing:
                    med_id = existing["id"]
                    updates: Dict[str, Any] = {
                        "price": selling_price,
                        "purchase_price": purchase_price,
                    }
                    if it.barcode and not existing.get("barcode"):
                        updates["barcode"] = it.barcode
                    if it.manufacturer:
                        updates["manufacturer"] = it.manufacturer
                    await _db.medicines.update_one({"id": med_id}, {"$set": updates})
                    updated_count += 1
                    is_new = False
                else:
                    med_id = str(uuid.uuid4())
                    await _db.medicines.insert_one({
                        "id": med_id,
                        "pharmacy_id": pharmacy_id,
                        "name": name,
                        "barcode": it.barcode,
                        "quantity": 0, "stock": 0,
                        "price": selling_price,
                        "purchase_price": purchase_price,
                        "expiry_date": expiry_iso,
                        "manufacturer": it.manufacturer,
                        "strength": it.strength,
                        "dosage_form": it.dosage_form,
                        "created_at": _now_iso(),
                    })
                    new_count += 1
                    is_new = True

                batch = await _batches.create_batch(
                    pharmacy_id, med_id, purchase_price, qty,
                    expiry_date=expiry_iso,
                )
                if it.batch_number:
                    await _db.medicine_batches.update_one(
                        {"id": batch["id"]},
                        {"$set": {"batch_number": it.batch_number}},
                    )

                new_total = await _batches.get_total_stock(pharmacy_id, med_id)
                await _db.medicines.update_one(
                    {"id": med_id},
                    {"$set": {"quantity": new_total, "stock": new_total}},
                )
                await _batches.refresh_medicine_expiry(pharmacy_id, med_id)

                succeeded.append({
                    "row": idx + 1, "name": name, "medicine_id": med_id,
                    "batch_id": batch["id"], "is_new": is_new,
                })
            except Exception as e:
                errors.append({"row": idx + 1, "name": (it.name or "")[:80],
                                "error": str(e)[:200]})

        return {
            "imported": len(succeeded),
            "new": new_count,
            "updated": updated_count,
            "failed": len(errors),
            "errors": errors,
            "succeeded": succeeded,
        }
