"""
AI-powered Supplier Catalog Import.

Pipeline:
1. Decode upload (PDF or image base64) -> list of base64 page images
2. For each page image -> Gemini 3 Flash extracts structured medicine entries (JSON)
3. Normalize names (Arabic-aware) -> dedupe within batch
4. For each item:
   - check `catalog_corrections` (learning) -> direct map
   - else fuzzy-match against supplier's existing products + canonical names
   - if confidence ambiguous -> ask Gemini "are these the same drug?" (switching)
5. Persist `import_items` with confidence + match_status (auto / needs_review)
"""
from __future__ import annotations

import base64
import io
import json
import logging
import os
import re
import uuid
from datetime import datetime, timezone
from typing import Any

import pypdfium2 as pdfium
from openpyxl import load_workbook, Workbook
from PIL import Image
from rapidfuzz import fuzz, process as rfprocess

logger = logging.getLogger(__name__)

EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")

# ---- Arabic / English normalization ----
_DIACRITICS = re.compile(r"[\u064B-\u0652\u0670\u0640]")


def normalize_text(s: str) -> str:
    s = (s or "").strip().lower()
    s = _DIACRITICS.sub("", s)
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    s = s.replace("ى", "ي").replace("ئ", "ي").replace("ؤ", "و")
    s = s.replace("ة", "ه")
    # Common pharma synonyms
    s = re.sub(r"\bgms?\b", "g", s)       # gm/gms -> g
    s = re.sub(r"\bgrams?\b", "g", s)
    s = re.sub(r"\bمل\b", "ml", s)
    s = re.sub(r"\bملغم?\b", "mg", s)
    s = re.sub(r"\bملي\s*غرام\b", "mg", s)
    s = re.sub(r"\bغرام\b", "g", s)
    s = re.sub(r"\btab(let)?s?\b", "tab", s)
    s = re.sub(r"\bcap(sule)?s?\b", "cap", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def canonical_key(name: str, strength: str | None = None,
                  dosage_form: str | None = None) -> str:
    parts = [normalize_text(name)]
    if strength:
        parts.append(normalize_text(strength))
    if dosage_form:
        parts.append(normalize_text(dosage_form))
    return " ".join(p for p in parts if p)


# ---- Excel structured parsing ----
# Header alias map: detected header (lowercased, trimmed) -> normalized field
HEADER_ALIASES: dict[str, str] = {
    # Name
    "name": "name", "product_name": "name", "product name": "name",
    "medicine": "name", "medicine_name": "name", "medicine name": "name",
    "drug": "name", "drug_name": "name",
    "اسم": "name", "اسم_الدواء": "name", "اسم الدواء": "name",
    "اسم_المنتج": "name", "اسم المنتج": "name", "المنتج": "name",
    "الدواء": "name", "الصنف": "name",
    # Price
    "price": "price", "unit_price": "price", "unit price": "price",
    "cost": "price", "rate": "price",
    "السعر": "price", "سعر": "price", "السعر_بالدينار": "price",
    "ثمن": "price",
    # Quantity
    "quantity": "quantity", "qty": "quantity", "stock": "quantity",
    "available": "quantity", "in_stock": "quantity", "in stock": "quantity",
    "amount": "quantity",
    "الكمية": "quantity", "المخزون": "quantity", "المتوفر": "quantity",
    "متاح": "quantity", "كمية": "quantity",
    # Strength / Dosage
    "strength": "strength", "dose": "strength", "dosage": "strength",
    "التركيز": "strength", "تركيز": "strength", "الجرعة": "strength",
    # Form
    "form": "dosage_form", "dosage_form": "dosage_form", "type": "dosage_form",
    "الشكل": "dosage_form", "نوع": "dosage_form",
    # Manufacturer
    "manufacturer": "manufacturer", "company": "manufacturer", "brand": "manufacturer",
    "الشركة": "manufacturer", "المصنع": "manufacturer", "الماركة": "manufacturer",
    # Category
    "category": "category", "group": "category",
    "الفئة": "category", "التصنيف": "category", "المجموعة": "category",
    # Delivery time
    "delivery_time": "delivery_time", "delivery time": "delivery_time",
    "وقت_التوصيل": "delivery_time", "وقت التوصيل": "delivery_time",
    # Expiry
    "expiry": "expiry_date", "expiry_date": "expiry_date", "expiration": "expiry_date",
    "exp": "expiry_date", "exp_date": "expiry_date",
    "تاريخ_الانتهاء": "expiry_date", "تاريخ الانتهاء": "expiry_date",
    "الانتهاء": "expiry_date",
}


def _norm_header(h: Any) -> str:
    s = str(h or "").strip().lower()
    s = _DIACRITICS.sub("", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _to_number(v: Any) -> float:
    """Best-effort numeric coercion. Strips currency, commas, Arabic digits."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return 0.0
    s = str(v).strip()
    # Arabic-Indic digits to ASCII
    arab_to_ascii = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    s = s.translate(arab_to_ascii)
    # Strip non-numeric except dot/minus
    s = re.sub(r"[^0-9\.\-]", "", s)
    if not s or s in (".", "-", "-.", ".-"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_excel_structured(file_b64: str) -> tuple[list[dict], dict]:
    """
    Try to read an .xlsx as a structured table.
    Returns (items, meta) where:
      - items: list of normalized dicts {name, strength, dosage_form, manufacturer, price, quantity, category, expiry_date}
      - meta: {sheet, columns_detected, total_rows, parsed_rows, structured_ok}
    `structured_ok` is False if no name+price columns can be confidently detected -> caller should fallback to AI.
    """
    raw = base64.b64decode(file_b64)
    bio = io.BytesIO(raw)
    wb = load_workbook(bio, read_only=True, data_only=True)
    meta: dict = {"sheets": wb.sheetnames, "structured_ok": False}
    # Use the first non-empty sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Find header row: first row with >=2 non-empty cells, scan first 8 rows
        header_row_idx = -1
        for i, r in enumerate(rows[:8]):
            non_empty = sum(1 for c in r if c not in (None, ""))
            if non_empty >= 2:
                header_row_idx = i
                break
        if header_row_idx == -1:
            continue
        header = rows[header_row_idx]
        # Map columns
        col_map: dict[int, str] = {}
        for idx, h in enumerate(header):
            key = _norm_header(h)
            if not key:
                continue
            field = HEADER_ALIASES.get(key)
            if not field:
                # Try fuzzy key match
                cand = list(HEADER_ALIASES.keys())
                best = rfprocess.extractOne(key, cand, scorer=fuzz.ratio)
                if best and best[1] >= 88:
                    field = HEADER_ALIASES[best[0]]
            if field:
                col_map[idx] = field
        # Need at least name + price for "structured_ok"
        fields = set(col_map.values())
        meta["sheet"] = sheet_name
        meta["columns_detected"] = {idx: f for idx, f in col_map.items()}
        meta["total_rows"] = max(0, len(rows) - header_row_idx - 1)
        if "name" not in fields or "price" not in fields:
            return [], meta
        items: list[dict] = []
        for r in rows[header_row_idx + 1:]:
            if not r or all(c in (None, "") for c in r):
                continue
            it: dict = {"name": "", "strength": None, "dosage_form": None,
                       "manufacturer": None, "price": 0.0, "quantity": 0,
                       "category": None, "expiry_date": None}
            for idx, field in col_map.items():
                if idx >= len(r):
                    continue
                val = r[idx]
                if field in ("price",):
                    it[field] = _to_number(val)
                elif field == "quantity":
                    it[field] = int(_to_number(val))
                elif field == "expiry_date":
                    it[field] = str(val).strip() if val not in (None, "") else None
                else:
                    it[field] = (str(val).strip() if val not in (None, "") else None)
            name = (it.get("name") or "").strip()
            price = float(it.get("price") or 0)
            if not name:
                continue
            # Validation: require name + price > 0
            if price <= 0:
                continue
            items.append(it)
        meta["parsed_rows"] = len(items)
        meta["structured_ok"] = len(items) > 0
        return items, meta
    return [], meta


def excel_to_text_dump(file_b64: str, max_rows: int = 200) -> str:
    """Convert first sheet of Excel to a text dump for AI fallback."""
    raw = base64.b64decode(file_b64)
    bio = io.BytesIO(raw)
    wb = load_workbook(bio, read_only=True, data_only=True)
    out_lines: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out_lines.append(f"=== Sheet: {sheet_name} ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                out_lines.append(f"... (truncated at {max_rows} rows)")
                break
            cells = [str(c) if c is not None else "" for c in row]
            out_lines.append("\t".join(cells))
        if out_lines:
            break  # only first non-empty sheet
    return "\n".join(out_lines)


def build_excel_template() -> bytes:
    """Build a sample .xlsx template suppliers can download."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"
    headers = ["product_name", "price", "quantity", "category", "strength", "dosage_form", "manufacturer", "expiry_date"]
    ws.append(headers)
    sample = [
        ["Paracetamol 500mg", 1500, 100, "مسكنات", "500mg", "tab", "GSK", "2027-12"],
        ["Amoxicillin 250mg", 2500, 50, "مضادات حيوية", "250mg", "cap", "Pfizer", "2026-08"],
        ["Vitamin C 1000mg", 3000, 200, "فيتامينات", "1000mg", "tab", "Bayer", "2028-05"],
    ]
    for row in sample:
        ws.append(row)
    # Set basic column widths
    widths = [22, 10, 12, 18, 12, 12, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---- PDF/Image decoding ----
def decode_to_page_images(file_b64: str, file_type: str) -> list[str]:
    """Return list of base64 JPEG images (one per page for PDF, one for image)."""
    raw = base64.b64decode(file_b64)
    pages: list[str] = []
    ft = (file_type or "").lower()
    if ft in ("pdf", "application/pdf"):
        pdf = pdfium.PdfDocument(raw)
        for i in range(min(len(pdf), 12)):  # safety cap: max 12 pages per upload
            page = pdf[i]
            pil_image = page.render(scale=2.0).to_pil()
            buf = io.BytesIO()
            pil_image.convert("RGB").save(buf, format="JPEG", quality=85)
            pages.append(base64.b64encode(buf.getvalue()).decode("ascii"))
    else:
        # Image: re-encode as JPEG to ensure compatibility
        img = Image.open(io.BytesIO(raw)).convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        pages.append(base64.b64encode(buf.getvalue()).decode("ascii"))
    return pages


# ---- Gemini extraction ----
EXTRACTION_PROMPT = """أنت محلل خبير لقوائم أسعار صيدلانية (price lists / catalogs).
استخرج كل دواء مذكور في الصورة وأرجع JSON ARRAY فقط (بدون أي شرح).
لكل دواء أرجع object بالحقول التالية:
{
  "name": "اسم الدواء بدون التركيز",
  "strength": "التركيز (مثل 500mg, 1g, 100ml) أو null",
  "dosage_form": "الشكل (tab, cap, syrup, injection, cream) أو null",
  "manufacturer": "الشركة الصانعة أو null",
  "price": رقم السعر بالدينار العراقي (أو 0),
  "quantity": رقم الكمية المتاحة (أو 0)
}
قواعد:
- إذا كان الجدول يحوي عمود "Available" أو "الكمية"، استخدمه لـ quantity
- إذا كان السعر بعملة أخرى، اتركه كما هو رقمياً
- لا تخمّن قيماً غير ظاهرة (استخدم null أو 0)
- أرجع JSON فقط، يبدأ بـ [ وينتهي بـ ]
"""


async def extract_items_from_image(image_b64: str) -> list[dict]:
    """Call Gemini 3 Flash on a single page image, return parsed list of items."""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"catalog-{uuid.uuid4()}",
            system_message="أنت مساعد متخصص في استخراج بيانات الأدوية من قوائم الأسعار.",
        ).with_model("gemini", "gemini-3-flash-preview")

        msg = UserMessage(text=EXTRACTION_PROMPT, file_contents=[ImageContent(image_base64=image_b64)])
        resp = await chat.send_message(msg)
        text = (resp or "").strip()
        # Strip code fences if present
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        # Find first [ to last ]
        start = text.find("[")
        end = text.rfind("]")
        if start == -1 or end == -1:
            return []
        items = json.loads(text[start:end + 1])
        if not isinstance(items, list):
            return []
        cleaned: list[dict] = []
        for it in items:
            if not isinstance(it, dict):
                continue
            name = (it.get("name") or "").strip()
            if not name:
                continue
            cleaned.append({
                "name": name,
                "strength": (it.get("strength") or None) or None,
                "dosage_form": (it.get("dosage_form") or None) or None,
                "manufacturer": (it.get("manufacturer") or None) or None,
                "price": float(it.get("price") or 0) if str(it.get("price") or "").replace(".", "", 1).replace("-", "", 1).isdigit() else 0.0,
                "quantity": int(it.get("quantity") or 0) if str(it.get("quantity") or "").isdigit() else 0,
            })
        return cleaned
    except Exception:
        logger.exception("extraction failed")
        return []


async def gemini_extract_from_text(text_dump: str) -> list[dict]:
    """Fallback: Send raw text dump (e.g., Excel rows) to Gemini for structured extraction."""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage

        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"catalog-text-{uuid.uuid4()}",
            system_message="أنت مساعد متخصص في استخراج بيانات الأدوية من الجداول.",
        ).with_model("gemini", "gemini-3-flash-preview")

        prompt = (
            "النص أدناه نسخة من ملف Excel/CSV لقائمة أسعار صيدلانية. "
            "استخرج كل دواء وأرجع JSON ARRAY فقط بنفس الحقول:\n"
            "{\"name\":..., \"strength\":..., \"dosage_form\":..., \"manufacturer\":..., \"price\":..., \"quantity\":...}\n"
            "قواعد: السعر بالدينار العراقي رقمياً، استخدم 0 إذا غير معروف، null للنصوص غير الواضحة.\n\n"
            f"=== المحتوى ===\n{text_dump[:8000]}\n=== انتهى ==="
        )
        resp = await chat.send_message(UserMessage(text=prompt))
        text = (resp or "").strip()
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        start = text.find("[")
        end = text.rfind("]")
        if start == -1 or end == -1:
            return []
        items = json.loads(text[start:end + 1])
        if not isinstance(items, list):
            return []
        cleaned: list[dict] = []
        for it in items:
            if not isinstance(it, dict):
                continue
            name = (it.get("name") or "").strip()
            if not name:
                continue
            cleaned.append({
                "name": name,
                "strength": (it.get("strength") or None) or None,
                "dosage_form": (it.get("dosage_form") or None) or None,
                "manufacturer": (it.get("manufacturer") or None) or None,
                "price": float(it.get("price") or 0) if str(it.get("price") or "").replace(".", "", 1).replace("-", "", 1).isdigit() else 0.0,
                "quantity": int(it.get("quantity") or 0) if str(it.get("quantity") or "").isdigit() else 0,
            })
        return cleaned
    except Exception:
        logger.exception("text extraction failed")
        return []


# ---- Smart matching ----
async def gemini_validate_match(text_a: str, text_b: str) -> float:
    """Ask Gemini if two pharmaceutical names refer to the same product. Returns confidence 0..1."""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage

        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"match-{uuid.uuid4()}",
            system_message="أنت خبير صيدلة. قارن أسماء الأدوية وأرجع JSON فقط.",
        ).with_model("gemini", "gemini-3-flash-preview")

        prompt = (
            f'هل هذان نفس الدواء؟\nA: "{text_a}"\nB: "{text_b}"\n'
            'أرجع JSON فقط: {"same": true|false, "confidence": 0.0-1.0}'
        )
        resp = await chat.send_message(UserMessage(text=prompt))
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", (resp or "").strip())
        m = re.search(r"\{[^}]+\}", text)
        if not m:
            return 0.5
        data = json.loads(m.group(0))
        same = bool(data.get("same"))
        conf = float(data.get("confidence") or 0)
        return conf if same else 1.0 - conf
    except Exception:
        return 0.5


async def smart_match(
    extracted_key: str,
    candidates: list[str],
    db_corrections: dict[str, str] | None = None,
) -> tuple[str | None, float]:
    """
    Returns (best_canonical_or_None, confidence 0..1).
    1. Check learned corrections (perfect override).
    2. RapidFuzz token_sort against candidates.
    3. If 70 < score < 90 -> Gemini re-validation (the "switching" layer).
    """
    if db_corrections and extracted_key in db_corrections:
        return db_corrections[extracted_key], 1.0
    if not candidates:
        return None, 0.0
    best = rfprocess.extractOne(extracted_key, candidates, scorer=fuzz.token_set_ratio)
    if not best:
        return None, 0.0
    match_str, score, _ = best
    conf = score / 100.0
    if conf >= 0.90:
        return match_str, conf
    if conf < 0.55:
        return None, conf  # too weak, treat as new
    # Ambiguous zone -> Gemini validation
    gemini_conf = await gemini_validate_match(extracted_key, match_str)
    if gemini_conf >= 0.7:
        return match_str, gemini_conf
    return None, max(conf, gemini_conf)


# ---- Main job processor ----
async def process_import_job(db, job_id: str) -> None:
    """Background task: read job, extract, match, persist items, update status."""
    job = await db.import_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        return
    try:
        await db.import_jobs.update_one({"id": job_id}, {"$set": {"status": "processing", "progress": 5}})

        ft = (job.get("file_type") or "").lower()
        is_excel = ("spreadsheetml" in ft or "excel" in ft or "xlsx" in ft
                    or (job.get("filename") or "").lower().endswith((".xlsx", ".xls", ".xlsm")))

        all_extracted: list[dict] = []
        method = "unknown"
        meta_extra: dict = {}

        if is_excel:
            # Hybrid: structured first, then Gemini fallback
            try:
                items, meta_extra = parse_excel_structured(job["file_b64"])
                if meta_extra.get("structured_ok") and items:
                    method = "excel_structured"
                    all_extracted = items
                    await db.import_jobs.update_one({"id": job_id},
                                                    {"$set": {"progress": 70}})
                else:
                    method = "excel_ai_fallback"
                    text_dump = excel_to_text_dump(job["file_b64"])
                    await db.import_jobs.update_one({"id": job_id},
                                                    {"$set": {"progress": 30}})
                    all_extracted = await gemini_extract_from_text(text_dump)
                    await db.import_jobs.update_one({"id": job_id},
                                                    {"$set": {"progress": 70}})
            except Exception as e:
                logger.exception("excel parse failed; trying AI fallback")
                method = "excel_ai_fallback"
                try:
                    text_dump = excel_to_text_dump(job["file_b64"])
                    all_extracted = await gemini_extract_from_text(text_dump)
                except Exception:
                    raise e
        else:
            method = "image_ai" if "image" in ft else "pdf_ai"
            page_images = decode_to_page_images(job["file_b64"], job["file_type"])
            await db.import_jobs.update_one({"id": job_id},
                                            {"$set": {"progress": 15, "page_count": len(page_images)}})
            for idx, img_b64 in enumerate(page_images):
                items = await extract_items_from_image(img_b64)
                all_extracted.extend(items)
                done = 15 + int(((idx + 1) / max(1, len(page_images))) * 60)
                await db.import_jobs.update_one({"id": job_id}, {"$set": {"progress": done}})

        # Dedupe within batch by canonical key
        dedup: dict[str, dict] = {}
        for it in all_extracted:
            ck = canonical_key(it["name"], it.get("strength"), it.get("dosage_form"))
            if not ck:
                continue
            if ck in dedup:
                # merge: keep cheaper price, sum quantity
                prev = dedup[ck]
                if it.get("price") and (not prev.get("price") or it["price"] < prev["price"]):
                    prev["price"] = it["price"]
                prev["quantity"] = (prev.get("quantity") or 0) + (it.get("quantity") or 0)
            else:
                dedup[ck] = {**it, "_canonical": ck}

        # Build candidate pool: this supplier's existing products + corrections
        existing = await db.supplier_products.find(
            {"supplier_id": job["supplier_id"]}, {"_id": 0, "name": 1}
        ).to_list(2000)
        candidates = list({normalize_text(p["name"]) for p in existing if p.get("name")})

        # Load supplier-specific corrections as dict
        corr_docs = await db.catalog_corrections.find(
            {"supplier_id": job["supplier_id"]}, {"_id": 0}
        ).to_list(5000)
        corrections = {c["original_key"]: c["corrected_name"] for c in corr_docs}

        items_to_insert: list[dict] = []
        review_count = 0
        rejected_invalid = 0
        for ck, it in dedup.items():
            # Validation: require name + price > 0
            if not it.get("name") or float(it.get("price") or 0) <= 0:
                rejected_invalid += 1
                continue
            match, conf = await smart_match(ck, candidates, corrections)
            status_ = "auto" if conf >= 0.90 else "needs_review"
            if status_ == "needs_review":
                review_count += 1
            items_to_insert.append({
                "id": str(uuid.uuid4()),
                "job_id": job_id,
                "supplier_id": job["supplier_id"],
                "raw_text": it["name"],
                "extracted": {
                    "name": it["name"],
                    "strength": it.get("strength"),
                    "dosage_form": it.get("dosage_form"),
                    "manufacturer": it.get("manufacturer"),
                    "price": it.get("price") or 0,
                    "quantity": it.get("quantity") or 0,
                    "category": it.get("category"),
                    "expiry_date": it.get("expiry_date"),
                },
                "canonical_key": ck,
                "suggested_canonical_name": match,
                "match_confidence": round(conf, 3),
                "match_status": status_,
                "approved_name": match if status_ == "auto" else None,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })

        if items_to_insert:
            await db.import_items.insert_many([dict(d) for d in items_to_insert])

        await db.import_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "status": "review",
                "progress": 100,
                "total_items": len(items_to_insert),
                "items_to_review": review_count,
                "rejected_invalid": rejected_invalid,
                "extraction_method": method,
                "extraction_meta": meta_extra,
                "completed_at": datetime.now(timezone.utc).isoformat(),
                # Drop the heavy file payload now to keep DB lean
                "file_b64": None,
            }},
        )
    except Exception as e:
        logger.exception("import job failed")
        await db.import_jobs.update_one(
            {"id": job_id},
            {"$set": {"status": "failed", "error": str(e)[:500],
                      "completed_at": datetime.now(timezone.utc).isoformat(),
                      "file_b64": None}},
        )
